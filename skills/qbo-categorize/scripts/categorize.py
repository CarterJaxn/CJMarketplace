#!/usr/bin/env python3
"""
L2M LLC — QBO transaction categorizer + classer.

Loads the knowledge base (QBO_KnowledgeBase_v2.json) and assigns every
transaction a GL category (account) AND a class.

CATEGORY precedence:  bank-account override -> exact vendor rule -> description keyword -> (review)
CLASS precedence:     account-default L2M  -> vendor fixed class -> bank-account store class -> (review)

Big shared vendors (Amazon, Amex, ADP...) carry class "By bank account (store)",
meaning their class is decided by whichever store's bank account paid them.

Usage:
    python3 categorize.py --kb QBO_KnowledgeBase_v2.json --in transactions.csv --out-dir .
    python3 categorize.py --selftest          # runs built-in sample transactions

Input transactions (CSV or XLSX). Column names are auto-detected (case-insensitive):
    date         : date | txn date | transaction date | posted
    vendor       : vendor | name | payee | description | merchant
    amount       : amount | amt | debit/credit
    bank_account : bank account | account | account name | source (the feed/account it came from)

Outputs (written to --out-dir):
    categorized.csv  -> import-ready rows that the rules handled confidently
    review.csv       -> rows needing a human decision (then feed corrections back into the KB)
    run_summary.txt  -> counts + coverage
"""
import argparse, csv, json, os, sys, datetime
from collections import Counter

FIXED_CLASSES = {"L2M Services", "ResaleAI", "Cool Springs", "Murfreesboro", "W Nashville"}

def norm(s):
    return str(s).strip().lower() if s is not None else ""

def acct_prefix(acct):
    """First token of an account string, e.g. '1101 80100 Main CSprings' -> '1101'."""
    if not acct:
        return ""
    return str(acct).strip().split()[0]

# ---------- knowledge base ----------
def load_kb(path):
    with open(path) as f:
        kb = json.load(f)
    kb.setdefault("vendor_rules", {})
    kb.setdefault("account_class_rules", {})
    kb.setdefault("bank_account_class", {})
    kb.setdefault("bank_account_category_overrides", [])
    kb.setdefault("description_patterns", [])
    # index bank-account class by numeric prefix for robust matching
    kb["_bank_class_by_prefix"] = {acct_prefix(k): v for k, v in kb["bank_account_class"].items()}
    # index overrides by (account-prefix, vendor-norm)
    ov = {}
    for o in kb["bank_account_category_overrides"]:
        ov[(acct_prefix(o.get("bank_account", "")), norm(o.get("vendor", "")))] = o.get("category", "")
    kb["_overrides"] = ov
    # normalize description patterns into (keyword, category) list
    pats = []
    for p in kb["description_patterns"]:
        for kw in str(p.get("keywords", "")).split(","):
            kw = kw.strip().lower()
            if kw:
                pats.append((kw, p.get("category", "")))
    kb["_patterns"] = pats
    kb["_skip"] = [s.strip().lower() for s in kb.get("skip_patterns", []) if s and s.strip()]
    return kb

# ---------- core rule engine ----------
def categorize_one(txn, kb):
    """Return dict with category, class, confidence, basis, needs_review, review_reason."""
    vendor_raw = txn.get("vendor", "")
    vn = norm(vendor_raw)
    bank = txn.get("bank_account", "")
    pfx = acct_prefix(bank)

    # ---- SKIP: structural transfers / payroll funding that QBO's native bank rules
    # already post at 100%. We recognize and skip these so they don't pollute review.
    skiptext = (vn + " " + norm(txn.get("description_extra", ""))).strip()
    for sk in kb.get("_skip", []):
        if sk and sk in skiptext:
            return {"category": "", "class": "", "cat_confidence": "", "category_basis": "",
                    "class_basis": "", "needs_review": False, "skipped": True,
                    "skip_reason": "structural/QBO-owned (" + sk + ")", "review_reason": ""}

    category = ""
    cat_basis = ""
    cat_conf = ""
    review_reason = []

    # ---- CATEGORY ----
    if (pfx, vn) in kb["_overrides"]:
        category = kb["_overrides"][(pfx, vn)]
        cat_basis = "bank-account override"
        cat_conf = "HIGH"
    elif vn and vn in kb["vendor_rules"]:
        r = kb["vendor_rules"][vn]
        category = r.get("category", "")
        cat_basis = "vendor rule"
        cat_conf = r.get("cat_confidence", "") or "MED"
    else:
        # description keyword (contains)
        desc = norm(txn.get("vendor", "")) + " " + norm(txn.get("description_extra", ""))
        for kw, cat in kb["_patterns"]:
            if kw in desc:
                category = cat
                cat_basis = "description keyword"
                cat_conf = "MED"
                break

    if not category:
        review_reason.append("no category rule")

    # ---- CLASS ----
    cls = ""
    cls_basis = ""
    if category and category in kb["account_class_rules"]:
        cls = kb["account_class_rules"][category]            # -> L2M Services
        cls_basis = "account-default L2M"
    else:
        rule = kb["vendor_rules"].get(vn)
        vendor_cls = rule.get("class", "") if rule else ""
        if vendor_cls in FIXED_CLASSES:
            cls = vendor_cls
            cls_basis = "vendor rule"
        else:
            # "By bank account (store)" or unknown -> use the paying account's class
            if pfx in kb["_bank_class_by_prefix"]:
                cls = kb["_bank_class_by_prefix"][pfx]
                cls_basis = "bank-account class"
            else:
                cls_basis = ""

    if not cls:
        review_reason.append("no class (unknown bank account)")
    if cat_conf == "LOW":
        review_reason.append("low-confidence category")

    return {
        "category": category,
        "class": cls,
        "cat_confidence": cat_conf,
        "category_basis": cat_basis,
        "class_basis": cls_basis,
        "needs_review": bool(review_reason),
        "review_reason": "; ".join(review_reason),
        "skipped": False,
        "skip_reason": "",
    }

# ---------- input loading ----------
CO="""date vendor amount bank_account""".split()
ALIASES = {
    "date": ["date", "txn date", "transaction date", "posted", "post date"],
    "vendor": ["vendor", "name", "payee", "description", "merchant", "memo"],
    "amount": ["amount", "amt", "value"],
    "bank_account": ["bank account", "bank_account", "account", "account name", "source", "feed"],
    "description_extra": ["full bank description", "bank description", "raw description", "description_extra"],
}

def map_columns(headers):
    hl = {h.lower().strip(): h for h in headers}
    out = {}
    for key, opts in ALIASES.items():
        for o in opts:
            if o in hl:
                out[key] = hl[o]
                break
    return out

def load_transactions(path):
    rows = []
    if path.lower().endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        data = list(ws.iter_rows(values_only=True))
        headers = [str(c) if c is not None else "" for c in data[0]]
        cmap = map_columns(headers)
        for r in data[1:]:
            d = dict(zip(headers, r))
            rows.append({k: d.get(cmap.get(k, ""), "") for k in ALIASES})
    else:
        with open(path, newline="") as f:
            rd = csv.DictReader(f)
            cmap = map_columns(rd.fieldnames or [])
            for d in rd:
                rows.append({k: d.get(cmap.get(k, ""), "") for k in ALIASES})
    return rows

# ---------- self-test ----------
SAMPLE = [
    {"date": "2026-06-01", "vendor": "AMAZON",          "amount": -120.00, "bank_account": "1121 80414 Main WBridge"},
    {"date": "2026-06-01", "vendor": "AMAZON",          "amount": -64.00,  "bank_account": "1101 80100 Main CSprings"},
    {"date": "2026-06-02", "vendor": "Tesla",           "amount": -812.00, "bank_account": "1130 ResaleAI Main"},
    {"date": "2026-06-02", "vendor": "Jackson Miller",  "amount": -5000.0, "bank_account": "1121 80414 Main WBridge"},
    {"date": "2026-06-03", "vendor": "SquareSpace",     "amount": -24.59,  "bank_account": "1130 ResaleAI Main"},
    {"date": "2026-06-03", "vendor": "Snapchat",        "amount": -300.00, "bank_account": "1111 80246 Main M'boro"},
    {"date": "2026-06-04", "vendor": "Dropbox",         "amount": -19.99,  "bank_account": "1121 80414 Main WBridge"},
    {"date": "2026-06-04", "vendor": "greko subs",      "amount": -42.10,  "bank_account": "1101 80100 Main CSprings"},  # keyword
    {"date": "2026-06-05", "vendor": "Some New Vendor LLC", "amount": -88.0, "bank_account": "1101 80100 Main CSprings"}, # review
]

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--kb", default="QBO_KnowledgeBase_v2.json")
    ap.add_argument("--in", dest="infile")
    ap.add_argument("--out-dir", default=".")
    ap.add_argument("--selftest", action="store_true")
    args = ap.parse_args()

    kb = load_kb(args.kb)
    if args.selftest:
        txns = SAMPLE
    elif args.infile:
        txns = load_transactions(args.infile)
    else:
        print("Provide --in <file> or --selftest", file=sys.stderr); sys.exit(2)

    os.makedirs(args.out_dir, exist_ok=True)
    cat_path = os.path.join(args.out_dir, "categorized.csv")
    rev_path = os.path.join(args.out_dir, "review.csv")
    skip_path = os.path.join(args.out_dir, "skipped.csv")
    fields = ["date", "vendor", "amount", "bank_account", "category", "class",
              "cat_confidence", "category_basis", "class_basis", "review_reason"]
    skip_fields = ["date", "vendor", "amount", "bank_account", "skip_reason"]
    n_cat = n_rev = n_skip = 0
    cls_dist = Counter()
    with open(cat_path, "w", newline="") as cf, open(rev_path, "w", newline="") as rf, open(skip_path, "w", newline="") as sf2:
        cw = csv.DictWriter(cf, fieldnames=fields); cw.writeheader()
        rw = csv.DictWriter(rf, fieldnames=fields); rw.writeheader()
        sw = csv.DictWriter(sf2, fieldnames=skip_fields); sw.writeheader()
        for t in txns:
            res = categorize_one(t, kb)
            base = {k: t.get(k, "") for k in ("date", "vendor", "amount", "bank_account")}
            if res.get("skipped"):
                sw.writerow({**base, "skip_reason": res.get("skip_reason", "")}); n_skip += 1
                continue
            row = {**base, **res}
            row = {k: row.get(k, "") for k in fields}
            if res["needs_review"]:
                rw.writerow(row); n_rev += 1
            else:
                cw.writerow(row); n_cat += 1
                cls_dist[res["class"]] += 1

    with open(os.path.join(args.out_dir, "run_summary.txt"), "w") as sf:
        total = n_cat + n_rev + n_skip
        actionable = n_cat + n_rev
        lines = [
            f"QBO categorization run — {datetime.datetime.now().isoformat(timespec='seconds')}",
            f"KB: {os.path.abspath(args.kb)}  (version {kb.get('version','?')}, built {kb.get('built','?')})",
            f"Transactions in: {total}",
            f"  Skipped (transfers/payroll — QBO native rules own these): {n_skip}",
            f"  Vendor transactions to handle: {actionable}",
            f"    Auto-categorized + classed: {n_cat} ({100*n_cat/actionable:.1f}%)" if actionable else "    none",
            f"    Sent to review:             {n_rev} ({100*n_rev/actionable:.1f}%)" if actionable else "",
            f"Class distribution (auto): {dict(cls_dist)}",
        ]
        sf.write("\n".join(lines) + "\n")
        print("\n".join(lines))
    print(f"\nWrote {cat_path}, {rev_path}, {skip_path}")

if __name__ == "__main__":
    main()
